"""Pre-carga del .xlsm diario de DimeroLab.

Toma la hoja `ProtocolosDigitales` (lo que cargaron los vetes en AppSheet y
pegaron en el Excel del día) y distribuye cada estudio a las planillas
correspondientes — equivalente al Sub Grabar() del VBA pero hecho en Python.

Planillas que llena:
- Planilla cobro          (1 fila por estudio, datos completos del paciente)
- Planilla Trabajo Hemograma   (componentes del hemograma o perfil)
- Planilla trabajo química     (componentes de la química o perfil)
- Planilla Orinas              (componentes de la orina)
- Planilla Coagulograma        (componentes del coagulograma)
- Planilla Serología           (1 fila por estudio de serología)
- Planilla Hemoparásitos       (1 fila por estudio)
- Planilla hematología         (1 fila por estudio)
- Planilla VetCheck            (1 fila completa por estudio, datos del paciente)
- Derivaciones                 (1 fila completa por derivación detectada)

Conservador: si no encuentra cómo distribuir un estudio (no es perfil ni
química conocida ni de los bloques especiales), igual escribe la fila en
Planilla cobro pero flaggea para revisión.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill
from unidecode import unidecode


DIAS_SEMANA = [
    "lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"
]
MESES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
]


# ─────────── helpers de texto ─────────────────────────────────────────────


def normalize(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = unidecode(s).lower()
    s = re.sub(r"\s+", " ", s)
    return s.rstrip(" .,;:")


def _sin_parens(s) -> str:
    return re.sub(r"\s*\([^)]*\)\s*", " ", str(s or "")).strip()


def _limpiar_nombre_estudio(s) -> str:
    """Limpia saltos de línea internos y whitespace duplicado del nombre
    del estudio. Los vetes a veces meten \\n al cargar en AppSheet."""
    if s in (None, ""):
        return ""
    txt = str(s).replace("\n", " ").replace("\r", " ").replace("\t", " ")
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


# Anotaciones de muestra que vienen pegadas al nombre del estudio.
# Cuando el vete escribe "Perfil completo LIPEMICO", la usuaria preserva
# el sufijo en la planilla cobro (es info importante para el técnico).
# Detectamos el sufijo, traducimos el estudio base al canónico y volvemos
# a pegar la anotación al final.
ANOTACIONES_MUESTRA = [
    "lipemico", "lipemica",
    "hemolizado", "hemolizada",
    "coagulado", "coagulada",
    "hg coagulado", "hemograma coagulado",
    "muestra escasa",
    "repeticion", "repetición",
    "otra vez",
    "sin cargo",
    "post dexametasona",
    "post estimulacion", "post estimulación",
    "post estimulacion acth", "post estimulación acth",
]


def _separar_anotacion_muestra(nombre: str) -> tuple[str, str]:
    """Devuelve (nombre_sin_anotacion, anotacion_original_con_caps).

    Si el nombre termina con una anotación conocida ('Perfil completo LIPEMICO'),
    la devuelve por separado preservando la capitalización original. Si no,
    devuelve (nombre, '').
    """
    if not nombre:
        return ("", "")
    n_normalizado = normalize(nombre)
    for anot in sorted(ANOTACIONES_MUESTRA, key=len, reverse=True):
        if n_normalizado.endswith(" " + anot):
            # Buscar el corte real en el texto original (preserva mayúsculas)
            # Suficiente: contar la longitud relativa al normalizado.
            # Como `normalize` solo baja a lower y deja whitespace simple,
            # los chars son equivalentes en cantidad. Hago búsqueda case-insensitive.
            largo_anot = len(anot)
            # Tomar los últimos chars del original (descartando trailing whitespace)
            stripped = nombre.rstrip()
            # Buscar la anotación al final del original ignorando case
            if stripped.lower().endswith(anot):
                base = stripped[: -largo_anot].rstrip()
                sufijo = stripped[-largo_anot:]  # como vino el vete
                return (base, sufijo)
    return (nombre, "")


def fecha_a_texto(valor) -> str:
    if valor is None or valor == "":
        return ""
    if isinstance(valor, (datetime, date)):
        d = valor
        return f"{d.day} {MESES[d.month - 1]} {DIAS_SEMANA[d.weekday()]}"
    return str(valor).strip()


# ─────────── aliases (traducción nombre vete → nombre canónico) ───────────


def cargar_aliases(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    return {
        normalize(k): v
        for k, v in raw.items()
        if not k.startswith("_") and isinstance(v, str)
    }


# ─────────── catálogo de químicas combinables (Albúmina, Urea, ...) ───────


@dataclass
class CatalogoQuimicas:
    """Químicas individuales que la usuaria agrupa como 'N química/s' en
    Planilla cobro. Una pedida (ej: Albúmina) sale como 1 química.
    Varias pedidas (ej: Urea + Creatinina) se cuentan como 2 químicas
    deduplicando por canónico.

    Equivalencias especiales: cuando un estudio pedido se descompone en
    varios canónicos (ej: 'Globulina' → Proteínas + Albúmina).
    """
    sinonimo_a_canonico: dict[str, str] = field(default_factory=dict)
    equivalencias_especiales: dict[str, list[str]] = field(default_factory=dict)

    def canonico_de(self, estudio: str) -> str | None:
        n = normalize(estudio)
        if n in self.sinonimo_a_canonico:
            return self.sinonimo_a_canonico[n]
        n_sp = normalize(_sin_parens(estudio))
        if n_sp and n_sp in self.sinonimo_a_canonico:
            return self.sinonimo_a_canonico[n_sp]
        return None

    def canonicos_equivalentes(self, estudio: str) -> list[str] | None:
        n = normalize(estudio)
        for k, v in self.equivalencias_especiales.items():
            if normalize(k) == n:
                return v
        return None


def cargar_catalogo_quimicas(path: Path) -> CatalogoQuimicas:
    if not path or not path.exists():
        return CatalogoQuimicas()
    raw = json.loads(path.read_text(encoding="utf-8"))
    sin_a_can: dict[str, str] = {}
    for q in raw.get("quimicas", []):
        canonico = q["canonico"]
        for s in q.get("sinonimos", []):
            sin_a_can[normalize(s)] = canonico
    equivs = {
        k: v
        for k, v in raw.get("equivalencias_especiales", {}).items()
        if not k.startswith("_")
    }
    return CatalogoQuimicas(sin_a_can, equivs)


# ─────────── derivables (estudio → destino) ───────────────────────────────


def cargar_derivables(path: Path) -> dict[str, str]:
    """{nombre_normalizado_del_estudio: destino_upper}

    Si un estudio cae acá, además de Planilla cobro se mete en Derivaciones.
    Si destino es 'EN EL LABO' es derivación interna (se imprime papel pero
    no sale del labo) — igual va a Derivaciones.
    """
    if not path or not path.exists():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    return {
        normalize(k): str(v).strip().upper()
        for k, v in raw.get("derivables", {}).items()
        if v
    }


def cargar_destinos(path: Path) -> dict[str, str]:
    """{destino_corto_upper: destino_completo_con_contacto}

    Ej: 'TCBA' → 'TCBA (DímeroLab: Francisco Zapata)'. Lo que va en el
    membrete impreso es el completo.
    """
    if not path or not path.exists():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    return {
        str(k).strip().upper(): str(v).strip()
        for k, v in raw.get("destinos", {}).items()
        if v
    }


# ─────────── catálogo del .xlsm ───────────────────────────────────────────


@dataclass
class Catalogo:
    """Estructura del .xlsm leída desde Perfiles.Estudios_con_planilla +
    los named ranges. Indexada por nombre normalizado del estudio."""

    # nombre_normalizado → ("perfil", índice_columna)
    perfiles: dict[str, int] = field(default_factory=dict)
    # nombre_normalizado → ("quimica", índice_columna)
    quimicas: dict[str, int] = field(default_factory=dict)
    # nombre_normalizado → ("coagulograma" | "orina" | ..., índice_columna, nombre_original)
    coagulogramas: dict[str, tuple[int, str]] = field(default_factory=dict)
    orinas: dict[str, tuple[int, str]] = field(default_factory=dict)
    serologia: dict[str, str] = field(default_factory=dict)
    hemoparasitos: dict[str, str] = field(default_factory=dict)
    vetcheck: dict[str, str] = field(default_factory=dict)
    hematologia: dict[str, str] = field(default_factory=dict)
    hemograma_canonico: str = "Hemograma completo"  # el nombre "Hemograma" lo trata especial
    # estudios_otros[i] = lista de nombres de estudios "otros" del perfil i
    estudios_otros_perfil: dict[int, list[str]] = field(default_factory=dict)
    # named ranges resueltos: nombre → lista de valores
    rangos: dict[str, list[str]] = field(default_factory=dict)
    # derivaciones: índice → texto de derivación (si != 0)
    deriva_perfil: dict[int, str] = field(default_factory=dict)
    derq_quimica: dict[int, str] = field(default_factory=dict)

    # Set ampliado de canónicos para "no flaggear" en Planilla cobro.
    canonicos_set: set[str] = field(default_factory=set)
    # Mapa nombre_normalizado → nombre_oficial_del_catálogo. Sirve para
    # devolver siempre la capitalización linda al hacer match (en lugar
    # del texto que tipea el vete).
    canonicos_nombre: dict[str, str] = field(default_factory=dict)


_FILA_HEMOGRAMA_TOKEN = 24
_FILA_QUIMICAS_INDIVIDUALES = 56
_FILA_COAGULOGRAMAS = 84
_FILA_ORINAS = 99
_FILA_DERIVA = 128
_FILA_DERQ = 138
_FILA_ESTUDIOS_OTROS_INICIO = 155
_FILA_SEROLOGIA = 168
_FILA_HEMOPARASITOS = 178
_FILA_VETCHECK = 189
_FILA_HEMATOLOGIA = 199


CANONICOS_FALLBACK = {
    "hemograma completo", "hemograma",
    "perfil completo",  # por si en alguna hoja no aparece exactamente
}


def _resolver_named_range(wb, nombre: str) -> list[str]:
    """Devuelve la lista de valores (no None) referenciada por un named range.
    Si no existe el rango o está vacío, devuelve []."""
    if nombre not in wb.defined_names:
        return []
    dn = wb.defined_names[nombre]
    valores: list[str] = []
    for sheet_name, coord in dn.destinations:
        try:
            ws = wb[sheet_name]
        except KeyError:
            continue
        target = ws[coord]
        # target puede ser Cell (1 celda), tuple de Cells (1 fila/columna)
        # o tuple de tuples (rango 2D). Normalizamos a iterable de Cells.
        if isinstance(target, openpyxl.cell.cell.Cell):
            celdas = [target]
        else:
            celdas = []
            for item in target:
                if isinstance(item, openpyxl.cell.cell.Cell):
                    celdas.append(item)
                else:
                    celdas.extend(item)
        for c in celdas:
            v = c.value
            if v not in (None, ""):
                valores.append(str(v).strip())
    return valores


def _resolver_celda(wb, nombre: str) -> str:
    """Devuelve el valor textual de la primera celda del named range, o ''."""
    vs = _resolver_named_range(wb, nombre)
    return vs[0] if vs else ""


def cargar_catalogo(wb, quimicas_path: Path | None = None) -> Catalogo:
    """Lee toda la estructura del .xlsm: perfiles, químicas y bloques especiales.

    Como los nombres de los rangos están normalizados a `Perfil1`, `Química1`,
    etc., y cada uno apunta a una columna distinta de Perfiles.Estudios_con_planilla,
    resolvemos cada named range una vez al cargar.
    """
    cat = Catalogo()
    if "Perfiles.Estudios_con_planilla" not in wb.sheetnames:
        return cat
    ws = wb["Perfiles.Estudios_con_planilla"]

    def _registrar_canonico(nombre: str) -> None:
        """Guarda el mapeo normalizado → nombre lindo del catálogo. Si ya
        estaba, no pisa (primer match gana)."""
        n = normalize(nombre)
        oficial = str(nombre).strip()
        if n and n not in cat.canonicos_nombre:
            cat.canonicos_nombre[n] = oficial
        n_sp = normalize(_sin_parens(nombre))
        if n_sp and n_sp not in cat.canonicos_nombre:
            cat.canonicos_nombre[n_sp] = oficial

    # Fila 2: nombres de perfiles (cols 1..20)
    for j in range(1, 25):
        v = ws.cell(row=2, column=j).value
        if v not in (None, "") and str(v).strip() not in ("0",):
            cat.perfiles[normalize(v)] = j
            _registrar_canonico(v)

    # Fila 56: nombres de químicas individuales
    for j in range(1, 25):
        v = ws.cell(row=_FILA_QUIMICAS_INDIVIDUALES, column=j).value
        if v not in (None, "") and str(v).strip() not in ("0",):
            cat.quimicas[normalize(v)] = j
            _registrar_canonico(v)

    def _bloque_simple(fila: int) -> dict[str, str]:
        out: dict[str, str] = {}
        for j in range(1, 25):
            v = ws.cell(row=fila, column=j).value
            if v not in (None, "") and str(v).strip() not in ("0",):
                out[normalize(v)] = str(v).strip()
                _registrar_canonico(v)
        return out

    # Bloques con named ranges asociados (coagulograma_i, orina_i)
    for j in range(1, 25):
        v = ws.cell(row=_FILA_COAGULOGRAMAS, column=j).value
        if v not in (None, "") and str(v).strip() not in ("0",):
            cat.coagulogramas[normalize(v)] = (j, str(v).strip())
            _registrar_canonico(v)
    for j in range(1, 25):
        v = ws.cell(row=_FILA_ORINAS, column=j).value
        if v not in (None, "") and str(v).strip() not in ("0",):
            cat.orinas[normalize(v)] = (j, str(v).strip())
            _registrar_canonico(v)

    cat.serologia = _bloque_simple(_FILA_SEROLOGIA)
    cat.hemoparasitos = _bloque_simple(_FILA_HEMOPARASITOS)
    cat.vetcheck = _bloque_simple(_FILA_VETCHECK)
    cat.hematologia = _bloque_simple(_FILA_HEMATOLOGIA)

    cat.hemograma_canonico = str(ws.cell(row=_FILA_HEMOGRAMA_TOKEN, column=1).value or "Hemograma completo").strip()

    # Estudios "otros" por perfil (fila 155+, columna del perfil)
    for j in cat.perfiles.values():
        otros: list[str] = []
        for r in range(_FILA_ESTUDIOS_OTROS_INICIO, _FILA_SEROLOGIA):
            v = ws.cell(row=r, column=j).value
            if v in (None, ""):
                break
            otros.append(str(v).strip())
        if otros:
            cat.estudios_otros_perfil[j] = otros

    # Derivaciones por perfil (fila 128) y por química (fila 138)
    for j in range(1, 25):
        v = ws.cell(row=_FILA_DERIVA, column=j).value
        if v not in (None, "", 0, "0"):
            cat.deriva_perfil[j] = str(v).strip()
        v = ws.cell(row=_FILA_DERQ, column=j).value
        if v not in (None, "", 0, "0"):
            cat.derq_quimica[j] = str(v).strip()

    # Resolver named ranges (componentes)
    for j in cat.perfiles.values():
        cat.rangos[f"Perfil{j}"] = _resolver_named_range(wb, f"Perfil{j}")
    for j in cat.quimicas.values():
        cat.rangos[f"Química{j}"] = _resolver_named_range(wb, f"Química{j}")
    for j_meta in cat.coagulogramas.values():
        j = j_meta[0]
        cat.rangos[f"Coagulograma{j}"] = _resolver_named_range(wb, f"Coagulograma{j}")
    for j_meta in cat.orinas.values():
        j = j_meta[0]
        cat.rangos[f"Orina{j}"] = _resolver_named_range(wb, f"Orina{j}")
    cat.rangos["Hemograma"] = _resolver_named_range(wb, "Hemograma")

    # perfil15 está como named range minúsculo (caso especial)
    if "perfil15" in wb.defined_names and "Perfil15" not in cat.rangos:
        cat.rangos["Perfil15"] = _resolver_named_range(wb, "perfil15")

    # ── Conjunto de canónicos para evitar flaggear "Hemograma completo" etc.
    cat.canonicos_set = set(CANONICOS_FALLBACK)
    cat.canonicos_set.update(cat.perfiles.keys())
    cat.canonicos_set.update(cat.quimicas.keys())
    cat.canonicos_set.update(cat.coagulogramas.keys())
    cat.canonicos_set.update(cat.orinas.keys())
    cat.canonicos_set.update(cat.serologia.keys())
    cat.canonicos_set.update(cat.hemoparasitos.keys())
    cat.canonicos_set.update(cat.vetcheck.keys())
    cat.canonicos_set.update(cat.hematologia.keys())
    cat.canonicos_set.add(normalize(cat.hemograma_canonico))
    # Variantes sin paréntesis
    cat.canonicos_set.update({normalize(_sin_parens(c)) for c in list(cat.canonicos_set)})

    # Sinónimos de químicas combinables (Albúmina, Glucemia, etc.)
    if quimicas_path and quimicas_path.exists():
        raw = json.loads(quimicas_path.read_text(encoding="utf-8"))
        for q in raw.get("quimicas", []):
            cat.canonicos_set.add(normalize(q.get("canonico", "")))
            for s in q.get("sinonimos", []):
                cat.canonicos_set.add(normalize(s))

    # Estudios "otros" de Todas_las_determinaciones (cols Otros estudios y Derivaciones)
    if "Todas_las_determinaciones" in wb.sheetnames:
        wsTd = wb["Todas_las_determinaciones"]
        for row in wsTd.iter_rows(min_row=2, values_only=True):
            for v in row:
                if v not in (None, ""):
                    cat.canonicos_set.add(normalize(v))
                    _registrar_canonico(v)

    cat.canonicos_set.discard("")
    return cat


# ─────────── lectura de ProtocolosDigitales ───────────────────────────────


@dataclass
class ProtocoloVete:
    fecha_txt: str
    codigo: int
    veterinaria: str
    nombre: str
    propietario: str
    especie: str
    raza: str
    edad: str
    sexo: str
    estudios: list[str] = field(default_factory=list)


def codigos_ya_cargados(wb) -> set[int]:
    """Lee la Planilla cobro y devuelve el set de códigos de protocolo que
    ya tienen al menos una fila cargada. Sirve para el modo incremental:
    solo procesar los protocolos nuevos."""
    if "Planilla cobro" not in wb.sheetnames:
        return set()
    ws = wb["Planilla cobro"]
    try:
        header_row = _find_header_row(ws, "CODIGO")
    except ValueError:
        return set()
    out: set[int] = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if len(row) < 2:
            continue
        cod = row[1]
        if cod in (None, ""):
            continue
        try:
            out.add(int(cod))
        except (TypeError, ValueError):
            continue
    return out


def cargar_protocolos(ws_protocolos) -> list[ProtocoloVete]:
    headers = [normalize(c.value) for c in ws_protocolos[1]]
    idx = {h: i for i, h in enumerate(headers) if h}
    col_fecha = idx.get("fecha")
    col_proto = idx.get("protocolo_ing")
    col_vet = idx.get("veterinaria_ing")
    col_paciente = idx.get("paciente")
    col_propietario = idx.get("propietario")
    col_especie = idx.get("especie")
    col_raza = idx.get("raza")
    col_edad = idx.get("edad")
    col_sexo = idx.get("sexo")
    cols_estudios = [
        idx.get(name)
        for name in ("p_estudio", "estudio_2", "estudio_3", "estudio_4", "estudio_5")
        if idx.get(name) is not None
    ]
    if col_proto is None:
        raise ValueError("ProtocolosDigitales: falta columna Protocolo_Ing")

    out: list[ProtocoloVete] = []
    for row in ws_protocolos.iter_rows(min_row=2, values_only=True):
        proto = row[col_proto] if col_proto is not None else None
        if proto in (None, ""):
            continue
        try:
            proto_int = int(proto)
        except (TypeError, ValueError):
            continue
        estudios = [
            _limpiar_nombre_estudio(row[c])
            for c in cols_estudios
            if row[c] not in (None, "")
        ]
        estudios = [e for e in estudios if e]
        if not estudios:
            continue
        out.append(
            ProtocoloVete(
                fecha_txt=fecha_a_texto(row[col_fecha]) if col_fecha is not None else "",
                codigo=proto_int,
                veterinaria=str(row[col_vet] or "") if col_vet is not None else "",
                nombre=str(row[col_paciente] or "") if col_paciente is not None else "",
                propietario=str(row[col_propietario] or "") if col_propietario is not None else "",
                especie=str(row[col_especie] or "") if col_especie is not None else "",
                raza=str(row[col_raza] or "") if col_raza is not None else "",
                edad=str(row[col_edad] or "") if col_edad is not None else "",
                sexo=str(row[col_sexo] or "") if col_sexo is not None else "",
                estudios=estudios,
            )
        )
    return out


# ─────────── escritura en planillas ────────────────────────────────────────


# Fill amarillo para filas de Planilla cobro cuyo estudio no se reconoció.
FILL_REVISAR = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")


def _find_header_row(ws, expected_token: str, max_scan: int = 10) -> int:
    token = normalize(expected_token)
    for r in range(1, min(ws.max_row, max_scan) + 1):
        for c in ws[r]:
            if normalize(c.value) == token:
                return r
    raise ValueError(f"No encontré header con '{expected_token}' en {ws.title}")


def _limpiar_desde_fila(ws, primera_fila: int) -> None:
    """Borra todas las filas desde primera_fila inclusive hasta el final.

    Usa delete_rows (operación nativa de openpyxl, mucho más rápida que iterar
    celda por celda — esas planillas pueden tener miles de filas acumuladas).
    """
    if ws.max_row < primera_fila:
        return
    cuantas = ws.max_row - primera_fila + 1
    if cuantas > 0:
        ws.delete_rows(primera_fila, cuantas)


@dataclass
class Distribucion:
    """Cuenta cuántas filas se escribieron en cada planilla."""
    cobro: int = 0
    hemograma: int = 0
    quimica: int = 0
    orinas: int = 0
    coagulograma: int = 0
    serologia: int = 0
    hemoparasitos: int = 0
    vetcheck: int = 0
    hematologia: int = 0
    derivaciones: int = 0


@dataclass
class FilaRevisar:
    codigo: int
    veterinaria: str
    estudio_original: str
    motivo: str  # "sin_alias_ni_canonico", "canonico_sin_distribucion"


def _normalize_eq(a: str, b: str) -> bool:
    """Compara dos nombres de estudio normalizados, también probando sin paréntesis."""
    if normalize(a) == normalize(b):
        return True
    a2 = normalize(_sin_parens(a))
    b2 = normalize(_sin_parens(b))
    return bool(a2) and bool(b2) and a2 == b2


class Distribuidor:
    """Acumula filas pendientes para cada planilla y al final las escribe en bloque.

    Esto evita modificar el worksheet mientras iteramos protocolos.
    """

    def __init__(
        self,
        wb,
        cat: Catalogo,
        catalogo_quimicas: CatalogoQuimicas | None = None,
        derivables: dict[str, str] | None = None,
        destinos_largos: dict[str, str] | None = None,
    ):
        self.wb = wb
        self.cat = cat
        self.catalogo_quimicas = catalogo_quimicas or CatalogoQuimicas()
        self.derivables = derivables or {}
        self.destinos_largos = destinos_largos or {}
        # Set para evitar duplicar entradas en Derivaciones (codigo, estudio_norm).
        self._derivaciones_vistas: set[tuple[int, str]] = set()
        # Contadores: destino_corto → cantidad de filas derivadas
        self.derivaciones_por_destino: dict[str, int] = {}
        # filas pendientes por planilla
        self.cobro: list[dict] = []
        self.hemograma: list[tuple[int, str]] = []  # (id, componente)
        self.quimica: list[tuple[int, str]] = []
        self.orinas: list[tuple[int, str]] = []
        self.coagulograma: list[tuple[int, str]] = []
        self.serologia: list[tuple[int, str]] = []
        self.hemoparasitos: list[tuple[int, str]] = []
        self.vetcheck: list[dict] = []  # fila completa
        self.hematologia: list[tuple[int, str]] = []
        self.derivaciones: list[dict] = []
        self.a_revisar: list[FilaRevisar] = []

    # ---- por planilla ----

    def _push_componentes(self, dest: list, codigo: int, nombre_rango: str) -> None:
        for componente in self.cat.rangos.get(nombre_rango, []):
            dest.append((codigo, componente))

    def _datos_paciente(self, p: ProtocoloVete, estudio: str) -> dict:
        return {
            "fecha": p.fecha_txt,
            "codigo": p.codigo,
            "veterinaria": p.veterinaria,
            "nombre": p.nombre,
            "propietario": p.propietario,
            "especie": p.especie,
            "raza": p.raza,
            "edad": p.edad,
            "sexo": p.sexo,
            "estudio": estudio,
        }

    def procesar_protocolo(self, p: ProtocoloVete, traducidos: list[tuple[str, str]]) -> None:
        """Procesa todos los estudios de un protocolo de una vez.

        Hace dos pasadas:
        1. Agrupa las químicas combinables (Albúmina, Urea, etc.) en una
           sola fila "N química/s" en Planilla cobro + N filas en química.
        2. Distribuye el resto de estudios según las reglas habituales.
        """
        canonicos_quimicas: list[str] = []  # mantengo orden, sin duplicados
        consumidos_quimicas: list[tuple[str, str]] = []
        no_quimicas: list[tuple[str, str]] = []

        for nombre, fuente in traducidos:
            # 1) equivalencia especial (Globulina → Proteínas + Albúmina)
            equiv = self.catalogo_quimicas.canonicos_equivalentes(nombre)
            if equiv:
                for c in equiv:
                    if c not in canonicos_quimicas:
                        canonicos_quimicas.append(c)
                consumidos_quimicas.append((nombre, fuente))
                continue
            # 2) sinónimo directo (Albúmina, Urea, Glucemia, ...)
            canon = self.catalogo_quimicas.canonico_de(nombre)
            if canon:
                if canon not in canonicos_quimicas:
                    canonicos_quimicas.append(canon)
                consumidos_quimicas.append((nombre, fuente))
                continue
            no_quimicas.append((nombre, fuente))

        # Si agarramos químicas combinables: una fila "N química/s" + N en Planilla química.
        if canonicos_quimicas:
            n = len(canonicos_quimicas)
            etiqueta = "1 química" if n == 1 else f"{n} químicas"
            self.cobro.append(self._datos_paciente(p, etiqueta))
            for canon in canonicos_quimicas:
                self.quimica.append((p.codigo, canon))

        # Resto de estudios → flujo normal.
        for nombre, fuente in no_quimicas:
            self._procesar_estudio_individual(p, nombre, fuente)

        # Derivaciones automáticas: para cada estudio (incluso los agrupados
        # en químicas), si está en la lista de derivables → fila en Derivaciones.
        for nombre, _ in traducidos:
            self._agregar_derivacion_si_corresponde(p, nombre)

    def _agregar_derivacion_si_corresponde(self, p: ProtocoloVete, estudio: str) -> None:
        n = normalize(estudio)
        n_sp = normalize(_sin_parens(estudio))
        destino = self.derivables.get(n) or self.derivables.get(n_sp)
        if not destino:
            return
        self._registrar_derivacion(p, estudio, destino=destino)

    def _resolver_destino(self, estudio: str) -> str | None:
        """Busca el destino para un estudio en derivables.json. Si no lo
        encuentra exacto, intenta dividir por comas/'+'/'y' y matchear partes
        — eso cubre casos como 'TSH, T4 total' (Deriva_i de perfil tiroideo)."""
        n = normalize(estudio)
        n_sp = normalize(_sin_parens(estudio))
        if n in self.derivables:
            return self.derivables[n]
        if n_sp in self.derivables:
            return self.derivables[n_sp]
        # Partir por comas / "+" / " y " para casos compuestos
        partes = re.split(r"\s*[,+]\s*|\s+y\s+", estudio)
        for parte in partes:
            np = normalize(parte)
            if np in self.derivables:
                return self.derivables[np]
            np_sp = normalize(_sin_parens(parte))
            if np_sp in self.derivables:
                return self.derivables[np_sp]
        return None

    def _registrar_derivacion(
        self,
        p: ProtocoloVete,
        estudio: str,
        destino: str | None = None,
    ) -> None:
        # Si el "estudio" es una cadena compuesta (ej: "TSH, T4 total")
        # y cada parte matchea individualmente en derivables, las registro
        # por separado para que cada una vaya a su destino.
        if destino is None:
            partes = [p_.strip() for p_ in re.split(r"\s*[,]\s*|\s+y\s+", estudio) if p_.strip()]
            if len(partes) > 1:
                destinos_partes = [self._resolver_destino(x) for x in partes]
                if any(d for d in destinos_partes):
                    for parte, d_parte in zip(partes, destinos_partes):
                        self._registrar_derivacion_simple(p, parte, d_parte)
                    return

        if destino is None:
            destino = self._resolver_destino(estudio)
        self._registrar_derivacion_simple(p, estudio, destino)

    def _registrar_derivacion_simple(
        self,
        p: ProtocoloVete,
        estudio: str,
        destino: str | None,
    ) -> None:
        n = normalize(estudio)
        key = (p.codigo, n)
        if key in self._derivaciones_vistas:
            return
        self._derivaciones_vistas.add(key)
        fila = self._datos_paciente(p, estudio)
        if destino:
            fila["destino"] = destino  # destino corto (TCBA, DIAP, etc.)
            fila["destino_largo"] = self.destinos_largos.get(destino, destino)
            self.derivaciones_por_destino[destino] = (
                self.derivaciones_por_destino.get(destino, 0) + 1
            )
        else:
            fila["destino_largo"] = "(definir)"
            self.derivaciones_por_destino["(definir)"] = (
                self.derivaciones_por_destino.get("(definir)", 0) + 1
            )
        self.derivaciones.append(fila)

    def _procesar_estudio_individual(self, p: ProtocoloVete, estudio_canon: str, fuente: str) -> None:
        """Distribuye un estudio según las reglas. Siempre escribe en Planilla cobro;
        además distribuye a planillas de trabajo si reconoce el tipo."""
        # 1) Siempre va a Planilla cobro
        fila_cobro = self._datos_paciente(p, estudio_canon)
        if fuente == "crudo":
            fila_cobro["_revisar"] = True
            self.a_revisar.append(FilaRevisar(
                codigo=p.codigo,
                veterinaria=p.veterinaria,
                estudio_original=estudio_canon,
                motivo="sin_alias_ni_canonico",
            ))
        self.cobro.append(fila_cobro)

        # 2) Buscar tipo y propagar
        n = normalize(estudio_canon)
        n_sp = normalize(_sin_parens(estudio_canon))

        distribuido = False

        # ── Perfil
        i_perfil = self.cat.perfiles.get(n) or self.cat.perfiles.get(n_sp)
        if i_perfil:
            distribuido = True
            self._push_componentes(self.quimica, p.codigo, f"Perfil{i_perfil}")
            self._push_componentes(self.hemograma, p.codigo, "Hemograma")
            if i_perfil in self.cat.deriva_perfil:
                deriv_estudio = self.cat.deriva_perfil[i_perfil]
                self._registrar_derivacion(p, deriv_estudio)
            # estudios otros del perfil
            for otro in self.cat.estudios_otros_perfil.get(i_perfil, []):
                self._distribuir_otro(p, otro)

        # ── Hemograma solo
        if not distribuido and (n == normalize(self.cat.hemograma_canonico) or n == "hemograma"):
            distribuido = True
            self._push_componentes(self.hemograma, p.codigo, "Hemograma")

        # ── Química individual
        if not distribuido:
            i_q = self.cat.quimicas.get(n) or self.cat.quimicas.get(n_sp)
            if i_q:
                distribuido = True
                self._push_componentes(self.quimica, p.codigo, f"Química{i_q}")
                if i_q in self.cat.derq_quimica:
                    deriv_estudio = self.cat.derq_quimica[i_q]
                    self._registrar_derivacion(p, deriv_estudio)

        # ── Coagulograma, Orina, Serología, Hemoparásitos, VetCheck, Hematología
        if not distribuido:
            distribuido = self._distribuir_bloque_especial(p, estudio_canon)

        # Si era canónico (matchea el catálogo amplio) pero no caímos en ninguno
        # de los buckets de distribución, queda solo en Planilla cobro. Esto es
        # esperable para estudios como "Cortisol Sérico" que no van a planilla
        # de trabajo. No flaggeamos como revisión.

    def _distribuir_otro(self, p: ProtocoloVete, otro: str) -> None:
        """Reaplica las reglas de bloque especial para un 'estudio otro' de perfil."""
        self._distribuir_bloque_especial(p, otro)

    def _distribuir_bloque_especial(self, p: ProtocoloVete, estudio: str) -> bool:
        n = normalize(estudio)
        n_sp = normalize(_sin_parens(estudio))

        # Coagulograma
        meta = self.cat.coagulogramas.get(n) or self.cat.coagulogramas.get(n_sp)
        if meta:
            i, _ = meta
            self._push_componentes(self.coagulograma, p.codigo, f"Coagulograma{i}")
            return True
        # Orina
        meta = self.cat.orinas.get(n) or self.cat.orinas.get(n_sp)
        if meta:
            i, _ = meta
            self._push_componentes(self.orinas, p.codigo, f"Orina{i}")
            return True
        # Serología
        if n in self.cat.serologia or n_sp in self.cat.serologia:
            nombre = self.cat.serologia.get(n) or self.cat.serologia.get(n_sp)
            self.serologia.append((p.codigo, nombre))
            return True
        # Hemoparásitos
        if n in self.cat.hemoparasitos or n_sp in self.cat.hemoparasitos:
            nombre = self.cat.hemoparasitos.get(n) or self.cat.hemoparasitos.get(n_sp)
            self.hemoparasitos.append((p.codigo, nombre))
            return True
        # VetCheck — fila completa
        if n in self.cat.vetcheck or n_sp in self.cat.vetcheck:
            nombre = self.cat.vetcheck.get(n) or self.cat.vetcheck.get(n_sp)
            self.vetcheck.append(self._datos_paciente(p, nombre))
            return True
        # Hematología
        if n in self.cat.hematologia or n_sp in self.cat.hematologia:
            nombre = self.cat.hematologia.get(n) or self.cat.hematologia.get(n_sp)
            self.hematologia.append((p.codigo, nombre))
            return True
        return False

    # ---- escritura final ----

    def volcar(self, incremental: bool = False) -> Distribucion:
        """Escribe todas las planillas.

        - Si incremental=False: borra el contenido previo de cada planilla
          y escribe todo desde cero.
        - Si incremental=True: NO borra nada; agrega filas al final de cada
          planilla (asume que las filas previas son de cargas anteriores
          que hay que preservar).
        """
        d = Distribucion()
        d.cobro = self._escribir_cobro_o_derivacion("Planilla cobro", self.cobro, header_token="CODIGO", incremental=incremental)
        d.derivaciones = self._escribir_cobro_o_derivacion("Derivaciones", self.derivaciones, header_token="CODIGO", incremental=incremental)
        d.hemograma = self._escribir_id_componente("Planilla Trabajo Hemograma", self.hemograma, incremental=incremental)
        d.quimica = self._escribir_id_componente("Planilla trabajo química", self.quimica, extra_col_d=True, incremental=incremental)
        d.orinas = self._escribir_id_componente("Planilla Orinas", self.orinas, incremental=incremental)
        d.coagulograma = self._escribir_id_componente("Planilla Coagulograma", self.coagulograma, incremental=incremental)
        d.serologia = self._escribir_id_componente("Planilla Serología", self.serologia, incremental=incremental)
        d.hemoparasitos = self._escribir_id_componente("Planilla Hemoparásitos", self.hemoparasitos, incremental=incremental)
        d.hematologia = self._escribir_id_componente("Planilla hematología", self.hematologia, incremental=incremental)
        d.vetcheck = self._escribir_vetcheck("Planilla VetCheck", self.vetcheck, incremental=incremental)
        return d

    def _primera_fila_libre(self, ws, desde: int) -> int:
        """Devuelve la primera fila vacía a partir de `desde`. Útil para modo
        incremental: encuentra dónde seguir escribiendo."""
        for r in range(desde, ws.max_row + 2):
            cell_a = ws.cell(row=r, column=1).value
            cell_b = ws.cell(row=r, column=2).value
            if cell_a in (None, "") and cell_b in (None, ""):
                return r
        return ws.max_row + 1

    def _escribir_cobro_o_derivacion(self, sheet: str, filas: list[dict], header_token: str, incremental: bool = False) -> int:
        if sheet not in self.wb.sheetnames:
            return 0
        ws = self.wb[sheet]
        header_row = _find_header_row(ws, header_token)
        if incremental:
            primera = self._primera_fila_libre(ws, header_row + 1)
        else:
            primera = header_row + 1
            _limpiar_desde_fila(ws, primera)
        # Headers de la planilla (mapping)
        headers = [normalize(c.value) for c in ws[header_row]]
        col = {h: i + 1 for i, h in enumerate(headers) if h}
        col_fecha = col.get("fecha")
        col_codigo = col.get("codigo")
        col_vet = col.get("veterinaria")
        col_nombre = col.get("nombre")
        col_prop = col.get("propietario")
        col_esp = col.get("especie")
        col_raza = col.get("raza")
        col_edad = col.get("edad")
        col_sexo = col.get("sexo")
        col_estudio = col.get("estudio solcitado") or col.get("estudio solicitado")

        for i, f in enumerate(filas):
            r = primera + i
            if col_fecha: ws.cell(row=r, column=col_fecha, value=f["fecha"])
            if col_codigo: ws.cell(row=r, column=col_codigo, value=f["codigo"])
            if col_vet: ws.cell(row=r, column=col_vet, value=f["veterinaria"])
            if col_nombre: ws.cell(row=r, column=col_nombre, value=f["nombre"])
            if col_prop: ws.cell(row=r, column=col_prop, value=f["propietario"])
            if col_esp: ws.cell(row=r, column=col_esp, value=f["especie"])
            if col_raza: ws.cell(row=r, column=col_raza, value=f["raza"])
            if col_edad: ws.cell(row=r, column=col_edad, value=f["edad"])
            if col_sexo: ws.cell(row=r, column=col_sexo, value=f["sexo"])
            if col_estudio: ws.cell(row=r, column=col_estudio, value=f["estudio"])
            if f.get("_revisar"):
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = FILL_REVISAR
        return len(filas)

    def _escribir_id_componente(
        self,
        sheet: str,
        filas: list[tuple[int, str]],
        extra_col_d: bool = False,
        incremental: bool = False,
    ) -> int:
        if sheet not in self.wb.sheetnames:
            return 0
        ws = self.wb[sheet]
        # Estas planillas tienen header en fila 1, datos desde fila 2.
        if incremental:
            primera = self._primera_fila_libre(ws, 2)
        else:
            primera = 2
            _limpiar_desde_fila(ws, primera)
        for i, (codigo, componente) in enumerate(filas):
            r = primera + i
            ws.cell(row=r, column=1, value=codigo)
            ws.cell(row=r, column=2, value=componente)
            if extra_col_d:
                ws.cell(row=r, column=4, value=f"{codigo}{componente}")
        return len(filas)

    def _escribir_vetcheck(self, sheet: str, filas: list[dict], incremental: bool = False) -> int:
        if sheet not in self.wb.sheetnames:
            return 0
        ws = self.wb[sheet]
        if incremental:
            primera = self._primera_fila_libre(ws, 2)
        else:
            primera = 2
            _limpiar_desde_fila(ws, primera)
        for i, f in enumerate(filas):
            r = primera + i
            ws.cell(row=r, column=1, value=f["fecha"])
            ws.cell(row=r, column=2, value=f["codigo"])
            ws.cell(row=r, column=3, value=f["veterinaria"])
            ws.cell(row=r, column=4, value=f["nombre"])
            ws.cell(row=r, column=5, value=f["propietario"])
            ws.cell(row=r, column=6, value=f["especie"])
            ws.cell(row=r, column=7, value=f["raza"])
            ws.cell(row=r, column=8, value=f["edad"])
            ws.cell(row=r, column=9, value=f["sexo"])
            ws.cell(row=r, column=10, value=f["estudio"])
        return len(filas)


# ─────────── traducción del nombre del estudio (alias/canónico) ─────────


def traducir_estudio(
    nombre_vete: str,
    aliases: dict[str, str],
    canonicos: set[str],
    canonicos_nombre: dict[str, str] | None = None,
) -> tuple[str, str]:
    """Devuelve (nombre_a_usar, fuente). fuente ∈ {'alias','canonico','crudo'}.

    Cuando matchea como canónico, devuelve el nombre lindo del catálogo
    (capitalización correcta) en lugar del texto tipeado por el vete.
    Preserva anotaciones de muestra al final ('LIPEMICO', 'HEMOLIZADO',
    'COAGULADO', 'REPETICION', etc.).
    """
    canonicos_nombre = canonicos_nombre or {}

    # Separar anotación de muestra si la hay (LIPEMICO, COAGULADO, etc.)
    base, anotacion = _separar_anotacion_muestra(nombre_vete)
    sufijo = f" {anotacion}" if anotacion else ""

    def _con_sufijo(s: str) -> str:
        return f"{s}{sufijo}"

    n = normalize(base)
    if n in aliases:
        return _con_sufijo(aliases[n]), "alias"
    if n in canonicos:
        return _con_sufijo(canonicos_nombre.get(n, base.strip())), "canonico"
    n_sp = normalize(_sin_parens(base))
    if n_sp and n_sp in aliases:
        return _con_sufijo(aliases[n_sp]), "alias"
    if n_sp and n_sp in canonicos:
        return _con_sufijo(canonicos_nombre.get(n_sp, base.strip())), "canonico"
    return str(nombre_vete).strip(), "crudo"


# ─────────── Templado V2 (membretes para imprimir derivaciones) ──────────


def agrupar_membretes(
    derivaciones: list[dict],
) -> list[dict]:
    """Agrupa derivaciones por (codigo, destino).

    Si un paciente tiene varios estudios al MISMO destino → 1 membrete con
    los estudios separados por ' + '. Si va a destinos distintos → 1 membrete
    por destino. La regla la decidió la usuaria 2026-05-28.

    Devuelve lista de dicts con keys: codigo, veterinaria, paciente,
    propietario, especie, raza, edad, sexo, estudio, destino.
    """
    grupos: dict[tuple[int, str], dict] = {}
    orden: list[tuple[int, str]] = []
    for d in derivaciones:
        destino_corto = d.get("destino") or "(definir)"
        destino_largo = d.get("destino_largo") or destino_corto
        key = (d["codigo"], destino_corto)
        if key not in grupos:
            grupos[key] = {
                "codigo": d["codigo"],
                "veterinaria": d["veterinaria"],
                "paciente": d["nombre"],
                "propietario": d["propietario"],
                "especie": d["especie"],
                "raza": d["raza"],
                "edad": d["edad"],
                "sexo": d["sexo"],
                "estudios": [d["estudio"]],
                "destino": destino_largo,  # el largo va al membrete
                "destino_corto": destino_corto,  # el corto para métricas/UI
            }
            orden.append(key)
        else:
            grupos[key]["estudios"].append(d["estudio"])
    out = []
    for k in orden:
        g = grupos[k]
        g["estudio"] = " + ".join(g.pop("estudios"))
        out.append(g)
    return out


# Mapeo columna Membretes → key del dict de membrete
_COL_MEMBRETE_KEY = {
    "A": "codigo", "B": "veterinaria", "C": "paciente", "D": "propietario",
    "E": "especie", "F": "raza", "G": "edad", "H": "sexo",
    "I": "estudio", "J": "muestra", "K": "destino",
}

# Regex para parsear fórmulas tipo '=Membretes!K2' o '=Membretes!$K$2'
_RX_FORMULA_MEMBRETES = re.compile(
    r"^=\s*Membretes!\$?([A-K])\$?(\d+)\s*$", re.IGNORECASE
)


def poblar_templado(template_bytes: bytes, membretes: list[dict]) -> bytes:
    """Abre el Templado V2, llena Membretes y preserva las fórmulas de Hoja1.

    Mantiene las fórmulas tipo `=Membretes!K2` en Hoja1 para que sigan
    funcionando cuando la usuaria edita Membretes. Pero como Excel no
    siempre recalcula al abrir archivos descargados de internet, post-
    procesamos el .xlsm e inyectamos el valor cacheado de cada fórmula
    directamente en el XML — así Hoja1 muestra los datos al abrir SIN
    depender de la recalculación, y si la usuaria edita una fila de
    Membretes, Excel detecta el cambio y la fórmula se recalcula.
    """
    wb = openpyxl.load_workbook(BytesIO(template_bytes), keep_vba=True, data_only=False)
    if "Membretes" not in wb.sheetnames:
        raise ValueError("El templado no tiene hoja 'Membretes'")

    # 1) Llenar Membretes (datos desde fila 2). Header en fila 1, cols:
    # A=Código, B=Veterinaria, C=Paciente, D=Propietario, E=Especie,
    # F=Raza, G=Edad, H=Sexo, I=Estudio, J=Muestra, K=Destino
    ws_m = wb["Membretes"]
    # Descombinar celdas en la zona de datos (filas 2 en adelante).
    # El template trae A27:K27 combinada con un valor residual; eso ocultaba
    # las columnas B-K del 26to membrete. Hacemos unmerge defensivo de
    # cualquier rango que toque la zona de datos.
    for rng in list(ws_m.merged_cells.ranges):
        if rng.min_row >= 2:
            ws_m.unmerge_cells(str(rng))
    _limpiar_desde_fila(ws_m, 2)
    for i, m in enumerate(membretes):
        r = 2 + i
        ws_m.cell(row=r, column=1, value=m["codigo"])
        ws_m.cell(row=r, column=2, value=m["veterinaria"])
        ws_m.cell(row=r, column=3, value=m["paciente"])
        ws_m.cell(row=r, column=4, value=m["propietario"])
        ws_m.cell(row=r, column=5, value=m["especie"])
        ws_m.cell(row=r, column=6, value=m["raza"])
        ws_m.cell(row=r, column=7, value=m["edad"])
        ws_m.cell(row=r, column=8, value=m["sexo"])
        ws_m.cell(row=r, column=9, value=m["estudio"])
        # Columna J (Muestra) en blanco — la completa la usuaria.
        ws_m.cell(row=r, column=11, value=m["destino"])

    # 2) Forzar a Excel a recalcular al abrir (por si el cached value falla).
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcCompleted = False

    # 3) Guardar a bytes — Hoja1 mantiene sus fórmulas originales intactas.
    buf = BytesIO()
    wb.save(buf)

    # 4) Post-procesar el .xlsm para inyectar cached values en Hoja1.
    return _inyectar_cached_values_hoja1(buf.getvalue(), membretes)


def _inyectar_cached_values_hoja1(xlsm_bytes: bytes, membretes: list[dict]) -> bytes:
    """Modifica el XML del .xlsm para que cada fórmula `=Membretes!XR` en
    Hoja1 tenga su `<v>VALOR</v>` cacheado. Así Excel muestra los datos
    al abrir sin necesidad de recalcular.
    """
    import zipfile
    from xml.etree import ElementTree as ET

    in_buf = BytesIO(xlsm_bytes)
    out_buf = BytesIO()

    with zipfile.ZipFile(in_buf, "r") as zin:
        # Encontrar qué archivo XML corresponde a Hoja1
        sheet_path = _resolver_xml_de_hoja(zin, "Hoja1")

        with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                data = zin.read(item)
                if sheet_path and item == sheet_path:
                    data = _patch_sheet_xml_con_cached_values(data, membretes)
                zout.writestr(item, data)

    return out_buf.getvalue()


def _resolver_xml_de_hoja(zin, nombre_hoja: str) -> str | None:
    """Dado un ZIP de .xlsm abierto, devuelve la ruta del XML que corresponde
    a la hoja con el nombre dado (ej: 'xl/worksheets/sheet1.xml')."""
    from xml.etree import ElementTree as ET

    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "rels": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    try:
        wb_xml = zin.read("xl/workbook.xml")
        rels_xml = zin.read("xl/_rels/workbook.xml.rels")
    except KeyError:
        return None

    # Buscar el rId de la hoja
    wb_root = ET.fromstring(wb_xml)
    rid = None
    for sheet in wb_root.iter(f"{{{ns['main']}}}sheet"):
        if sheet.attrib.get("name") == nombre_hoja:
            rid = sheet.attrib.get(f"{{{ns['r']}}}id")
            break
    if not rid:
        return None

    # Resolver el rId a un target
    rels_root = ET.fromstring(rels_xml)
    for rel in rels_root.iter(f"{{{ns['rels']}}}Relationship"):
        if rel.attrib.get("Id") == rid:
            target = rel.attrib.get("Target", "")
            # Normalizar a ruta dentro del ZIP. Target puede venir como:
            # - "/xl/worksheets/sheet1.xml" (absoluto desde root del package)
            # - "xl/worksheets/sheet1.xml" (con prefijo xl)
            # - "worksheets/sheet1.xml" (relativo a xl/)
            t = target.lstrip("/")
            if t.startswith("xl/"):
                return t
            return f"xl/{t}"
    return None


def _patch_sheet_xml_con_cached_values(xml_bytes: bytes, membretes: list[dict]) -> bytes:
    """Para cada `<c r="..."><f>=Membretes!XR</f></c>`, agrega `<v>VALOR</v>`
    con el valor calculado a partir de los datos de membretes."""

    text = xml_bytes.decode("utf-8")

    # Regex que matchea celdas con fórmula a Membretes. openpyxl escribe:
    #   <c r="A1" s="27"><f>Membretes!K2</f><v /></c>
    # Nota: la fórmula NO tiene "=" al inicio en el XML, y el <v /> viene
    # self-closing (sin valor) o como <v>...</v> si ya hay cached.
    pattern = re.compile(
        r'<c\b([^>]*?)>'              # 1: atributos completos del <c>
        r'\s*<f\b[^>]*>'              # apertura de <f>
        r'\s*=?\s*Membretes!\$?([A-K])\$?(\d+)\s*'  # 2: col letra, 3: fila
        r'</f>'                       # cierre de <f>
        r'(?:\s*<v\s*/>|\s*<v[^>]*>[^<]*</v>)?'  # opcional: <v/> o <v>...</v>
        r'\s*</c>',
        re.IGNORECASE,
    )

    def replace_fn(m):
        attrs = m.group(1)  # algo como ' r="A1" s="27"'
        col_letra = m.group(2).upper()
        fila_membrete = int(m.group(3))
        idx = fila_membrete - 2  # primer membrete es fila 2

        key = _COL_MEMBRETE_KEY.get(col_letra)
        if not key or not (0 <= idx < len(membretes)):
            return m.group(0)

        valor = membretes[idx].get(key, "")
        if valor in (None, ""):
            return m.group(0)

        # Limpiar atributos: quitar t="..." si está
        attrs_limpio = re.sub(r'\s+t="[^"]*"', '', attrs)

        if isinstance(valor, (int, float)):
            cached = str(valor)
            return (
                f'<c{attrs_limpio}>'
                f'<f>Membretes!{col_letra}{fila_membrete}</f>'
                f'<v>{cached}</v>'
                f'</c>'
            )
        else:
            cached = _xml_escape(str(valor))
            return (
                f'<c{attrs_limpio} t="str">'
                f'<f>Membretes!{col_letra}{fila_membrete}</f>'
                f'<v>{cached}</v>'
                f'</c>'
            )

    text = pattern.sub(replace_fn, text)
    return text.encode("utf-8")


def _xml_escape(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


# ─────────── entrada principal ────────────────────────────────────────────


@dataclass
class Resultado:
    xlsm_bytes: bytes
    distribucion: Distribucion
    a_revisar: list[FilaRevisar] = field(default_factory=list)
    metricas: dict = field(default_factory=dict)
    templado_bytes: bytes | None = None
    membretes: list[dict] = field(default_factory=list)


def correr(
    xlsm_bytes: bytes,
    aliases_path: Path,
    quimicas_path: Path | None = None,
    derivables_path: Path | None = None,
    templado_path: Path | None = None,
    modo: str = "completo",
) -> Resultado:
    """`modo`:
    - 'completo' (default): borra el contenido previo de las planillas y
      las regenera desde cero.
    - 'incremental': preserva lo que ya está cargado, solo procesa los
      protocolos cuyo código no esté en Planilla cobro. Útil cuando llega
      una segunda tanda de protocolos en el mismo día.
    """
    wb = openpyxl.load_workbook(BytesIO(xlsm_bytes), keep_vba=True, data_only=False)

    if "ProtocolosDigitales" not in wb.sheetnames:
        raise ValueError("El archivo no tiene la hoja 'ProtocolosDigitales'")
    if "Planilla cobro" not in wb.sheetnames:
        raise ValueError("El archivo no tiene la hoja 'Planilla cobro'")

    if quimicas_path is None:
        quimicas_path = aliases_path.parent / "quimicas_combinables.json"
    if derivables_path is None:
        derivables_path = aliases_path.parent / "derivables.json"
    if templado_path is None:
        templado_path = aliases_path.parent / "templado_derivaciones.xlsm"
    destinos_path = aliases_path.parent / "destinos.json"

    aliases = cargar_aliases(aliases_path)
    cat = cargar_catalogo(wb, quimicas_path)
    catalogo_q = cargar_catalogo_quimicas(quimicas_path)
    derivables = cargar_derivables(derivables_path)
    destinos_largos = cargar_destinos(destinos_path)
    protocolos = cargar_protocolos(wb["ProtocolosDigitales"])

    # Modo incremental: solo procesar protocolos que no estén ya cargados.
    incremental = (modo == "incremental")
    ya_cargados: set[int] = set()
    n_omitidos = 0
    if incremental:
        ya_cargados = codigos_ya_cargados(wb)
        antes = len(protocolos)
        protocolos = [p for p in protocolos if p.codigo not in ya_cargados]
        n_omitidos = antes - len(protocolos)

    dist = Distribuidor(wb, cat, catalogo_q, derivables, destinos_largos)
    n_alias = n_canon = n_crudo = 0
    for p in protocolos:
        traducidos: list[tuple[str, str]] = []
        for estudio_vete in p.estudios:
            traducido, fuente = traducir_estudio(
                estudio_vete, aliases, cat.canonicos_set, cat.canonicos_nombre
            )
            traducidos.append((traducido, fuente))
            if fuente == "alias": n_alias += 1
            elif fuente == "canonico": n_canon += 1
            else: n_crudo += 1
        dist.procesar_protocolo(p, traducidos)

    resumen = dist.volcar(incremental=incremental)

    buffer = BytesIO()
    wb.save(buffer)

    metricas = {
        "modo": modo,
        "protocolos_ya_cargados": len(ya_cargados),
        "protocolos_omitidos_por_existir": n_omitidos,
        "protocolos_unicos": len({p.codigo for p in protocolos}),
        "estudios_totales": sum(len(p.estudios) for p in protocolos),
        "estudios_alias": n_alias,
        "estudios_canonico": n_canon,
        "estudios_a_revisar": n_crudo,
        "filas_planilla_cobro": resumen.cobro,
        "filas_hemograma": resumen.hemograma,
        "filas_quimica": resumen.quimica,
        "filas_orinas": resumen.orinas,
        "filas_coagulograma": resumen.coagulograma,
        "filas_serologia": resumen.serologia,
        "filas_hemoparasitos": resumen.hemoparasitos,
        "filas_vetcheck": resumen.vetcheck,
        "filas_hematologia": resumen.hematologia,
        "filas_derivaciones": resumen.derivaciones,
        "derivaciones_por_destino": dist.derivaciones_por_destino,
        "aliases_cargados": len(aliases),
        "canonicos_cargados": len(cat.canonicos_set),
        "perfiles_cargados": len(cat.perfiles),
        "quimicas_individuales_cargadas": len(cat.quimicas),
        "derivables_cargados": len(derivables),
    }
    # Templado V2 con membretes para imprimir derivaciones.
    membretes = agrupar_membretes(dist.derivaciones)
    templado_bytes: bytes | None = None
    if templado_path and templado_path.exists() and membretes:
        try:
            with open(templado_path, "rb") as f:
                templado_bytes = poblar_templado(f.read(), membretes)
            metricas["membretes_generados"] = len(membretes)
        except Exception as e:
            metricas["templado_error"] = str(e)

    return Resultado(
        xlsm_bytes=buffer.getvalue(),
        distribucion=resumen,
        a_revisar=dist.a_revisar,
        metricas=metricas,
        templado_bytes=templado_bytes,
        membretes=membretes,
    )
